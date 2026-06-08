from openpyxl import load_workbook

def col_letter(n):
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result

wb = load_workbook('Score Sheet/Call Centre Audit Report - October 2024 (Recovered).xlsx')

# Check Oct 2024 sheet (main data)
ws_oct = wb['Oct 2024']
print("OCT 2024 SHEET (Main Data):")
print(f"Dimensions: {ws_oct.dimensions}")
print(f"Row 1 (headers): ", end="")
for col in range(1, 15):
    cell = ws_oct.cell(row=1, column=col)
    if cell.value:
        print(f"{cell.coordinate}={cell.value}", end=" | ")
print("\n")

# Team Performance detailed
ws_team = wb['Team performance']
print("TEAM PERFORMANCE - First 12 rows, columns A-N:")
for row in range(1, 13):
    row_data = []
    for col in range(1, 15):
        cell = ws_team.cell(row=row, column=col)
        val = cell.value
        if val is None:
            row_data.append("")
        else:
            if isinstance(val, float):
                row_data.append(f"{col_letter(col)}{row}:{val:.4f}")
            else:
                row_data.append(f"{col_letter(col)}{row}:{str(val)[:15]}")
    print(f"Row {row}: {' | '.join(row_data)}")

# Trend sheet detailed
ws_trend = wb['Trend']
print("\n\nTREND SHEET - All data (rows 1-6, cols A-R):")
for row in range(1, 7):
    row_data = []
    for col in range(1, 19):
        cell = ws_trend.cell(row=row, column=col)
        val = cell.value
        if val is None:
            row_data.append("")
        elif isinstance(val, float):
            row_data.append(f"{val:.4f}")
        else:
            row_data.append(str(val)[:20])
    print(f"Row {row}: {row_data}")

wb.close()
