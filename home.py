import streamlit as st
import os
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import glob
import io
import re
from extra_streamlit_components import CookieManager

st.set_page_config(page_title="QA Audit Platform", layout="wide")

# --- User Authentication and RBAC ---
import streamlit_authenticator as stauth

# Define credentials dictionary for streamlit-authenticator >=0.4.2
credentials = {
    "usernames": {
        "amanda.bio-marfo@nationwidemh.com": {
            "name": "Amanda Bio-Marfo",
            "password": stauth.Hasher.hash("123")
        },
        "suraiyatu.mohammed@nationwidemh.com": {
            "name": "Surayatu Mohammed",
            "password": stauth.Hasher.hash("456")
        },
        "edem.dzitrie@nationwidemh.com": {
            "name": "Edem Dzitrie",
            "password": stauth.Hasher.hash("789")
        }
    }
}

cookie_manager = CookieManager(key="qa_app_home_cookie_manager")

authenticator = stauth.Authenticate(
    credentials=credentials,
    cookie_name="qa_app_home",
    key="qa_app_home_auth",
    expiry_days=1,
    cookie_manager=cookie_manager
)

# Login widget
authenticator.login(location='main')

# Get values from session state
name = st.session_state.get("name")
authentication_status = st.session_state.get("authentication_status")
username = st.session_state.get("username")

def get_user_role(username):
    # Example static mapping; replace with DB lookup as needed
    role_map = {
        "amanda.bio-marfo@nationwidemh.com": "admin",
        "suraiyatu.mohammed@nationwidemh.com": "auditor",
        "edem.dzitrie@nationwidemh.com": "supervisor"
    }
    return role_map.get(username, None)

if authentication_status == False:
    st.error("Invalid credentials")
    st.stop()
elif authentication_status is None:
    st.warning("Please enter your username and password")
    st.stop()

# If we reach here, user is authenticated
st.sidebar.success(f"Welcome {name}!")

# Navigation menu
st.sidebar.title("📋 Navigation")
app_mode = st.sidebar.radio(
    "Select View:",
    ["🏠 Home", "🛠️ Tool", "📊 Dashboard"],
    key="app_mode"
)

# User role
user_role = get_user_role(username)
st.sidebar.markdown(f"**Role:** {user_role}")

# Logout button
if st.sidebar.button("🚪 Logout"):
    st.session_state.authentication_status = False
    st.rerun()

# ============================================
# HOME PAGE
# ============================================
if app_mode == "🏠 Home":
    st.title("🏠 QA Audit Platform")
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.info("### 📊 Dashboard\nView and analyze audit data with interactive visualizations and performance metrics.")
        if st.button("Go to Dashboard", key="btn_dashboard"):
            st.session_state.app_mode = "📊 Dashboard"
            st.rerun()
    
    with col2:
        st.info("### 🛠️ Tool\nManage and process audit calls with advanced tools and utilities.")
        if st.button("Go to Tool", key="btn_tool"):
            st.session_state.app_mode = "🛠️ Tool"
            st.rerun()
    
    with col3:
        st.info("### ℹ️ Platform Info\nAccess documentation and help resources.")
        st.write("Learn more about using the QA Audit Platform.")
    
    st.markdown("---")
    st.subheader("Quick Stats")
    
    try:
        audit_files = glob.glob("Audit_log_calls/audit_log_*.xlsx")
        if audit_files:
            dfs = []
            for file in audit_files:
                try:
                    df = pd.read_excel(file)
                    dfs.append(df)
                except:
                    pass
            
            if dfs:
                audit_data = pd.concat(dfs, ignore_index=True)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("📞 Total Calls", len(audit_data))
                with col2:
                    if 'Total Score' in audit_data.columns:
                        avg_score = pd.to_numeric(audit_data['Total Score'], errors='coerce').mean()
                        st.metric("⭐ Avg Score", f"{avg_score:.1f}" if not pd.isna(avg_score) else "N/A")
                with col3:
                    if 'Name of Call Centre Officer' in audit_data.columns:
                        agents = audit_data['Name of Call Centre Officer'].nunique()
                        st.metric("👥 Agents", agents)
                with col4:
                    if 'Date of interaction' in audit_data.columns:
                        dates = pd.to_datetime(audit_data['Date of interaction'], errors='coerce').dt.date.nunique()
                        st.metric("📅 Days", dates)
    except:
        st.info("No audit data available yet.")

# ============================================
# TOOL PAGE
# ============================================
elif app_mode == "🛠️ Tool":
    st.title("🛠️ QA Audio Call Selector")
    
    # Initialize session state for uploaded files
    if 'uploaded_files_folder' not in st.session_state:
        st.session_state.uploaded_files_folder = None
    
    # Choose input method
    input_method = st.radio("Choose input method:", ["📁 From Folder Path", "📤 Upload Files"], horizontal=True)
    
    folder_path = None
    
    if input_method == "📁 From Folder Path":
        # Folder input
        folder_path = st.text_input("📁 Enter or select folder path containing call recordings:")
    
    else:  # Upload Files
        # File uploader
        st.write("**Upload recorded call files (MP3, WAV, M4A)**")
        uploaded_files = st.file_uploader(
            "Choose audio files to upload",
            type=["mp3", "wav", "m4a"],
            accept_multiple_files=True,
            key="call_file_uploader"
        )
    
        if uploaded_files:
            # Create temporary directory for uploaded files if not already created
            if st.session_state.uploaded_files_folder is None:
                import tempfile
                st.session_state.uploaded_files_folder = tempfile.mkdtemp()
    
            temp_dir = st.session_state.uploaded_files_folder
    
            # Save uploaded files
            for uploaded_file in uploaded_files:
                file_path = os.path.join(temp_dir, uploaded_file.name)
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
    
            st.success(f"✅ {len(uploaded_files)} file(s) uploaded successfully!")
            folder_path = temp_dir
        else:
            st.info("No files uploaded yet. Please select audio files to proceed.")
            folder_path = None
    
    
    if folder_path and os.path.isdir(folder_path):
        files = [f for f in os.listdir(folder_path) if f.lower().endswith((".mp3", ".wav", ".m4a"))]
    
        if not files:
            st.warning("No audio files found in this folder.")
        else:
            st.success(f"Found {len(files)} recordings.")
    
            # Optional: extract agent name or date from filenames
            df = pd.DataFrame({"File_Name": files})
    
            # work from filename without extension
            df["Base"] = df["File_Name"].str.rsplit(".", n=1).str[0]
    
            # Agent: first segment before the first underscore
            df["Agent"] = df["Base"].str.split("_").str[0].str.strip("[]").fillna("Unknown").replace("", "Unknown")
    
            # extract date from the third segment (after second "_"), with common-format fallbacks
            def _extract_date_from_base(base):
                parts = base.split("_")
                seg = parts[2] if len(parts) > 2 else ""
                if not seg:
                    return "Unknown"
                m = re.search(r"(\d{4}-\d{2}-\d{2})", seg)          # YYYY-MM-DD
                if m:
                    return m.group(1)
                m = re.search(r"(\d{8})", seg)                    # YYYYMMDD
                if m:
                    s = m.group(1)
                    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
                m = re.search(r"(\d{2}-\d{2}-\d{4})", seg)        # DD-MM-YYYY
                if m:
                    d = m.group(1)
                    day, month, year = d.split("-")
                    return f"{year}-{month}-{day}"
                return "Unknown"
    
            df["Date"] = df["Base"].apply(_extract_date_from_base)
            df["Contact"] = df["File_Name"].apply(lambda x: os.path.join(folder_path, x))
    
            # Initialize all QA columns with exact names from audit report
            df["Id"] = range(1, len(df) + 1)
            df["Name"] = ""
            df["Date of interaction"] = df["Date"]
            df["AHT (seconds)"] = ""
            df["Agent"] = df["Name of Call Centre Officer"] = df["Agent"]
            df["Type of caller"] = ""
            df["Name of caller"] = ""
            df["Caller's phone number"] = ""
            df["Caller's organization/Name of facility"] = ""
            df["Language spoken"] = ""
            df["Purpose of call"] = ""
            df["Did CCO open the call using the appropriate greetings?"] = ""
            df["Was the CCO able to identify and verify the needs of the customer?"] = ""
            df["Was the CCO able to educate & inform the customer about the query/enquiry/request"] = ""
            df["Did the CCO ensure and confirm the necessary steps to query resolution?"] = ""
            df["Did the CCO accept responsibility for the query? (CAN DO)"] = ""
            df["Did the CCO speak clearly and fluently throughout the call?"] = ""
            df["Was the CCO professional?"] = ""
            df["Was the CCO able to communicate effectively through effective listening and troubleshooting?"] = ""
            df["Was the CCO polite & courteous?"] = ""
            df["Did the CCO show empathy? (introduce solution statement)"] = ""
            df["Did the CCO ask if you had any further needs?"] = ""
            df["Did the CCO end the call politely and professionally?"] = ""
            df["Total Score"] = 0
            df["Fatal"] = ""
            df["QA Score"] = 0.0
            df["Do you have any other comments you would like to share?"] = ""
    
            # Choose sampling type
            sampling_type = st.radio("🎯 Sampling Type", ["Pure Random", "Stratified by Agent"])
            sample_size = st.number_input("🔢 Number of Calls to Audit", min_value=1, value=5)
    
            if st.button("🎲 Select Random Calls"):
                selected = pd.DataFrame()
    
                if sampling_type == "Pure Random":
                    selected = df.sample(n=min(sample_size, len(df)))
    
                elif sampling_type == "Stratified by Agent":
                    agents = df["Agent"].unique()
                    total_calls = len(df)
                    samples_per_agent = {}
                    total_allocated = 0
                    for a in agents:
                        num = len(df[df["Agent"] == a])
                        proportion = num / total_calls if total_calls > 0 else 0
                        allocated = round(sample_size * proportion)
                        if allocated == 0 and num > 0:
                            allocated = 1
                        samples_per_agent[a] = min(allocated, num)
                        total_allocated += samples_per_agent[a]
    
                    # If total allocated is less than sample_size, add more to agents with remaining capacity
                    if total_allocated < sample_size:
                        remaining = sample_size - total_allocated
                        agents_sorted = sorted(agents, key=lambda a: len(df[df["Agent"] == a]) - samples_per_agent[a], reverse=True)
                        for a in agents_sorted:
                            can_add = len(df[df["Agent"] == a]) - samples_per_agent[a]
                            if can_add > 0:
                                add = min(remaining, can_add)
                                samples_per_agent[a] += add
                                remaining -= add
                                if remaining == 0:
                                    break
    
                    for a in agents:
                        subset = df[df["Agent"] == a]
                        selected = pd.concat([selected, subset.sample(n=samples_per_agent[a])])
    
                # Store selected in session state
                st.session_state.selected = selected.copy()
    
            # Display selections if available
            if 'selected' in st.session_state:
                selected = st.session_state.selected
                st.subheader("✅ Selected Calls for QA Audit")
    
                # Initialize session state for all fields if not present
                # Benchmark-aligned quality questions and their scores
                quality_questions = [
                    ("Did CCO open the call using the appropriate greetings?", 3),
                    ("Was the CCO able to identify and verify the needs of the customer?", 4),
                    ("Was the CCO able to educate & inform the customer about the query/enquiry/request", 15),
                    ("Did the CCO ensure and confirm the necessary steps to query resolution?", 5),
                    ("Did the CCO accept responsibility for the query? (CAN DO)", 5),
                    ("Did the CCO speak clearly and fluently throughout the call?", 5),
                    ("Was the CCO professional?", 15),
                    ("Was the CCO able to communicate effectively through effective listening and troubleshooting?", 15),
                    ("Was the CCO polite & courteous?", 10),
                    ("Did the CCO show empathy? (introduce solution statement)", 15),
                    ("Did the CCO ask if you had any further needs?", 5),
                    ("Did the CCO end the call politely and professionally?", 3)
                ]
    
                for i in selected.index:
                    # Metadata fields
                    if f"name_{i}" not in st.session_state:
                        st.session_state[f"name_{i}"] = selected.at[i, 'Name'] if selected.at[i, 'Name'] else ''
                    if f"aht_{i}" not in st.session_state:
                        st.session_state[f"aht_{i}"] = selected.at[i, 'AHT (seconds)'] if selected.at[i, 'AHT (seconds)'] else ''
                    if f"caller_type_{i}" not in st.session_state:
                        st.session_state[f"caller_type_{i}"] = selected.at[i, 'Type of caller'] if selected.at[i, 'Type of caller'] else ''
                    if f"caller_name_{i}" not in st.session_state:
                        st.session_state[f"caller_name_{i}"] = selected.at[i, 'Name of caller'] if selected.at[i, 'Name of caller'] else ''
                    if f"cc_officer's_name_{i}" not in st.session_state:
                        st.session_state[f"cc_officer's_name_{i}"] = selected.at[i, 'Name of Call Centre Officer'] if selected.at[i, 'Name of Call Centre Officer'] else ''
                    if f"caller_phone_{i}" not in st.session_state:
                        st.session_state[f"caller_phone_{i}"] = selected.at[i, "Caller's phone number"] if selected.at[i, "Caller's phone number"] else ''
                    if f"caller_org_{i}" not in st.session_state:
                        st.session_state[f"caller_org_{i}"] = selected.at[i, "Caller's organization/Name of facility"] if selected.at[i, "Caller's organization/Name of facility"] else ''
                    if f"language_{i}" not in st.session_state:
                        st.session_state[f"language_{i}"] = selected.at[i, 'Language spoken'] if selected.at[i, 'Language spoken'] else ''
                    if f"purpose_{i}" not in st.session_state:
                        st.session_state[f"purpose_{i}"] = selected.at[i, 'Purpose of call'] if selected.at[i, 'Purpose of call'] else ''
    
                    # Quality metrics
                    for q, _ in quality_questions:
                        if f"{q}_{i}" not in st.session_state:
                            st.session_state[f"{q}_{i}"] = selected.at[i, q] if q in selected.columns and selected.at[i, q] else ''
    
                    # Other fields
                    if f"fatal_{i}" not in st.session_state:
                        st.session_state[f"fatal_{i}"] = selected.at[i, 'Fatal'] if selected.at[i, 'Fatal'] else ''
                    if f"comments_{i}" not in st.session_state:
                        st.session_state[f"comments_{i}"] = selected.at[i, 'Do you have any other comments you would like to share?'] if selected.at[i, 'Do you have any other comments you would like to share?'] else ''
    
                # Collect widget values
                form_data = {}
                for i, row in selected.iterrows():
                    st.write(f"## 🎙️ Call {i+1}: {row['File_Name']}")
                    st.write(f"**Agent:** {row['Agent']} | **Date:** {row['Date']}")
                    st.audio(row["Contact"])
    
                    # Cancel button for Not Applicable (N/A)
                    if st.button(f"Cancel (N/A) for Call {i+1}", key=f"cancel_na_{i}"):
                        st.session_state[f"na_{i}"] = True
                    if f"na_{i}" in st.session_state and st.session_state[f"na_{i}"]:
                        st.info("This call has been marked as Not Applicable (N/A) and will be skipped.")
                        form_data[f"na_{i}"] = True
                        st.divider()
                        continue
                    else:
                        form_data[f"na_{i}"] = False
    
                    with st.expander("📋 Call Metadata & Caller Info", expanded=True):
                        col1, col2 = st.columns(2)
                        with col1:
                            form_data[f"name_{i}"] = st.text_input(f"Auditor's Name", value=st.session_state[f"name_{i}"], key=f"name_input_{i}")
                            form_data[f"aht_{i}"] = st.number_input(f"AHT (seconds)", value=int(st.session_state[f"aht_{i}"]) if st.session_state[f"aht_{i}"] else 0, key=f"aht_input_{i}")
                            form_data[f"cc_officer's_name_{i}"] = st.text_input(f"CC Officer's Name", value=st.session_state[f"cc_officer's_name_{i}"], key=f"cc_officer_name_input_{i}")
                            form_data[f"caller_name_{i}"] = st.text_input(f"Caller Name", value=st.session_state[f"caller_name_{i}"], key=f"caller_name_input_{i}")
                        with col2:
                            form_data[f"caller_phone_{i}"] = st.text_input(f"Caller Phone", value=st.session_state[f"caller_phone_{i}"], key=f"caller_phone_input_{i}")
                            form_data[f"caller_org_{i}"] = st.text_input(f"Caller Organization", value=st.session_state[f"caller_org_{i}"], key=f"caller_org_input_{i}")
                            form_data[f"language_{i}"] = st.selectbox(f"Language Spoken", ["", "English", "Twi", "Ga", "Other"], index=["", "English", "Twi", "Ga", "Other"].index(st.session_state[f"language_{i}"] if st.session_state[f"language_{i}"] else ""), key=f"language_input_{i}")
                            form_data[f"purpose_{i}"] = st.selectbox(f"Purpose of Call", ["", "Complaint", "Enquiry", "Request"], index=["", "Complaint", "Enquiry", "Request"].index(st.session_state[f"purpose_{i}"] if st.session_state[f"purpose_{i}"] else ""), key=f"purpose_input_{i}")
    
                    with st.expander("⭐ Quality Metrics (Rate as Yes/No)", expanded=True):
                        col1, col2 = st.columns(2)
                        for idx, (q, score) in enumerate(quality_questions):
                            label = f"{q} (Score: {score})"
                            if q == "Was the CCO able to identify and verify the needs of the customer?":
                                options = ["", "0", "3", "4", "NA"]
                            elif q == "Was the CCO able to educate & inform the customer about the query/enquiry/request":
                                options = ["", "0", "5", "10", "13", "15", "Fatal"]
                            elif q == "Did the CCO ensure and confirm the necessary steps to query resolution?":
                                options = ["", "0", "4", "5", "NA"]
                            elif q == "Did the CCO accept responsibility for the query? (CAN DO)":
                                options = ["", "0", "2", "5", "NA"]
                            elif q == "Did the CCO speak clearly and fluently throughout the call?":
                                options = ["", "0", "2", "5"]
                            elif q == "Was the CCO professional?":
                                options = ["", "0", "3", "6", "9", "12", "15"]
                            elif q == "Was the CCO able to communicate effectively through effective listening and troubleshooting?":
                                options = ["", "0", "5", "10", "14", "15"]
                            elif q == "Was the CCO polite & courteous?":
                                options = ["", "0", "5", "10", "Fatal"]
                            elif q == "Did the CCO show empathy? (introduce solution statement)":
                                options = ["", "0", "10", "13", "14", "15", "NA"]
                            else:
                                options = ["", "0", "1", "2", "3", "NA"]
                            idx_default = options.index(str(st.session_state[f"{q}_{i}"]) if st.session_state[f"{q}_{i}"] else "")
                            if idx % 2 == 0:
                                form_data[f"{q}_{i}"] = col1.selectbox(label, options, index=idx_default, key=f"{q}_input_{i}")
                            else:
                                form_data[f"{q}_{i}"] = col2.selectbox(label, options, index=idx_default, key=f"{q}_input_{i}")
    
                    with st.expander("📝 Additional Info", expanded=False):
                        form_data[f"comments_{i}"] = st.text_area(f"Comments", value=st.session_state[f"comments_{i}"], key=f"comments_input_{i}")
    
                    st.divider()
    
                # Save audit log button
                if st.button("💾 Save Audit Log"):
    
                    # Update selected from form data
                    selected_save = st.session_state.selected.copy()
    
                    for i in selected_save.index:
                        # If marked as N/A, set a flag and skip scoring/inputs
                        if form_data.get(f"na_{i}", False):
                            selected_save.at[i, 'Not Applicable'] = 'N/A'
                            selected_save.at[i, 'Total Score'] = 0
                            selected_save.at[i, 'QA Score'] = 0.0
                            selected_save.at[i, 'Do you have any other comments you would like to share?'] = form_data.get(f"comments_{i}", '')
                            continue
    
                        # Metadata
                        selected_save.at[i, 'Name'] = form_data.get(f"name_{i}", '')
                        selected_save.at[i, 'AHT (seconds)'] = form_data.get(f"aht_{i}", '')
                        selected_save.at[i, 'Type of caller'] = form_data.get(f"caller_type_{i}", '')
                        selected_save.at[i, 'Name of caller'] = form_data.get(f"caller_name_{i}", '')
                        selected_save.at[i, "Caller's phone number"] = form_data.get(f"caller_phone_{i}", '')
                        selected_save.at[i, "Caller's organization/Name of facility"] = form_data.get(f"caller_org_{i}", '')
                        selected_save.at[i, 'Language spoken'] = form_data.get(f"language_{i}", '')
                        selected_save.at[i, 'Purpose of call'] = form_data.get(f"purpose_{i}", '')
    
                        # Quality metrics
                        total_score = 0
                        for q, score in quality_questions:
                            answer = form_data.get(f"{q}_{i}", '')
                            selected_save.at[i, q] = answer
                            if answer == "Yes":
                                total_score += score
    
                        # Other fields
                        selected_save.at[i, 'Do you have any other comments you would like to share?'] = form_data.get(f"comments_{i}", '')
    
                        # Compute Total Score and QA Score
                        selected_save.at[i, 'Total Score'] = total_score
                        selected_save.at[i, 'QA Score'] = (total_score / 100) if total_score > 0 else 0.0
    
                    # Show current working directory for debugging
                    st.info(f"Current working directory: {os.getcwd()}")
                    st.write("Data to be saved:")
                    st.dataframe(selected_save)
    
                    try:
                        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        selected_save["Audit_Timestamp"] = timestamp
                        filename = f"Audit_log_calls/audit_log_{timestamp}.xlsx"
                        os.makedirs("Audit_log_calls", exist_ok=True)
                        selected_save.to_excel(filename, index=False)
                        st.success(f"✅ Audit log saved successfully! File: {filename}")
                    except Exception as e:
                        st.error(f"Failed to save audit log: {e}")
    
                    # Optional: display summary
                    st.write("**Audit Summary:**")
                    st.dataframe(selected_save[["Agent", "File_Name", "Date", "Total Score", "Do you have any other comments you would like to share?"]])
    
                    # QA Statistics
                    st.write("**QA Statistics:**")
                    if not selected_save.empty and selected_save['Total Score'].notna().any():
                        avg_score = selected_save['Total Score'].mean()
                        st.write(f"Overall Average Total Score: {avg_score:.2f} / 100")
                        # Ensure all columns are string type before groupby
                        stats_df = selected_save.copy()
                        stats_df['Agent'] = stats_df['Agent'].astype(str)
                        stats_df['Total Score'] = stats_df['Total Score'].astype(int)
                        agent_stats = stats_df.groupby('Agent')['Total Score'].agg(['mean', 'count']).reset_index()
                        agent_stats = agent_stats.rename(columns={'mean': 'Average Score', 'count': 'Call Count'})
                        st.write("Average Score and Call Count by Agent:")
                        st.dataframe(agent_stats)
                        # Bar chart
                        fig = px.bar(
                            agent_stats,
                            x='Agent',
                            y='Average Score',
                            title='Average Total Score by Agent',
                            labels={'Average Score': 'Average Score', 'Agent': 'Agent Name'},
                            color='Average Score',
                            color_continuous_scale='RdYlGn',
                            range_color=[0, 100]
                        )
                        fig.update_layout(
                            font=dict(family="Arial, sans-serif", size=14),
                            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
                            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
                            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
                        )
                        st.plotly_chart(fig)
                    else:
                        st.write("No QA scores available for statistics.")
                    st.write(f"Total Comments Provided: {selected_save['Do you have any other comments you would like to share?'].str.strip().ne('').sum()}")

# ============================================
# DASHBOARD PAGE
# ============================================
elif app_mode == "📊 Dashboard":
    st.title("📊 QA Audit Dashboard")
    
    # Load and combine all audit log files into audit_data
    audit_files = glob.glob("Audit_log_calls/audit_log_*.xlsx")
    if not audit_files:
        st.error("No audit log files found. Please ensure audit_log_*.xlsx files are present.")
        st.stop()
    
    dfs = []
    for file in audit_files:
        try:
            df = pd.read_excel(file)
            dfs.append(df)
        except Exception as e:
            st.warning(f"Could not read {file}: {e}")
    
    if not dfs:
        st.error("No valid audit log data loaded.")
        st.stop()
    
    audit_data = pd.concat(dfs, ignore_index=True)
    
    # Try to infer a consistent datetime format for 'Audit_Timestamp' to avoid the warning
    if 'Audit_Timestamp' in audit_data.columns:
        ts = pd.to_datetime(audit_data['Audit_Timestamp'], format="%d/%m/%Y %H:%M", errors='coerce')
    elif 'Date of interaction' in audit_data.columns:
        ts = pd.to_datetime(audit_data['Date of interaction'], format="%d/%m/%Y %H:%M", errors='coerce')
    else:
        ts = None
    
    date_range = None
    if ts is not None and ts.notna().any():
        min_dt = ts.min().date()
        max_dt = ts.max().date()
        date_range = st.sidebar.date_input(
            "📅 Date Range",
            value=(min_dt, max_dt),
            min_value=min_dt,
            max_value=max_dt
        )
    
    filtered_data = audit_data.copy()
    if ts is not None and isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        start_date, end_date = date_range
        if start_date is not None and end_date is not None:
            valid_ts = ts.notna()
            start_ts = pd.to_datetime(start_date)
            end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
            filtered_data = filtered_data[valid_ts & (ts >= start_ts) & (ts <= end_ts)]
    
    # Standardize agent column for consistency
    if 'Agent' not in filtered_data.columns and 'Name of Call Centre Officer' in filtered_data.columns:
        filtered_data['Agent'] = filtered_data['Name of Call Centre Officer']
    elif 'Agent' in filtered_data.columns and 'Name of Call Centre Officer' not in filtered_data.columns:
        filtered_data['Name of Call Centre Officer'] = filtered_data['Agent']
    
    # Agent filter
    agents = sorted(filtered_data['Agent'].dropna().unique())
    selected_agents = st.sidebar.multiselect("👥 Select Agents", agents, default=agents)
    if selected_agents:
        filtered_data = filtered_data[filtered_data['Agent'].isin(selected_agents)]
    
    # Ensure Total Score is numeric and comments are strings
    filtered_data = filtered_data.copy()
    if 'Total Score' in filtered_data.columns:
        filtered_data['Total Score'] = pd.to_numeric(filtered_data['Total Score'], errors='coerce')
    if 'QA Score' in filtered_data.columns:
        filtered_data['QA Score'] = pd.to_numeric(filtered_data['QA Score'], errors='coerce')
    if 'Do you have any other comments you would like to share?' in filtered_data.columns:
        filtered_data['Do you have any other comments you would like to share?'] = filtered_data['Do you have any other comments you would like to share?'].fillna('').astype(str)
    else:
        filtered_data['Do you have any other comments you would like to share?'] = ''
    
    # Score threshold (dynamic min/max to avoid RangeError)
    if 'Total Score' in filtered_data.columns and filtered_data['Total Score'].notna().any():
        min_score = int(filtered_data['Total Score'].min())
        max_score = int(filtered_data['Total Score'].max())
        if min_score == max_score:
            score_range = st.sidebar.slider("⭐ Total Score Range", min_score, min_score+1, (min_score, min_score+1))
        else:
            score_range = st.sidebar.slider("⭐ Total Score Range", min_score, max_score, (min_score, max_score))
        filtered_data = filtered_data[
            filtered_data['Total Score'].notna() &
            (filtered_data['Total Score'] >= score_range[0]) &
            (filtered_data['Total Score'] <= score_range[1])
        ]
    else:
        st.sidebar.info("No scores available for range filter.")
    
    st.sidebar.write(f"📌 Showing {len(filtered_data)} records")
    
    # Main metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📞 Total Calls Audited", len(filtered_data))
    
    with col2:
        avg_score = filtered_data['Total Score'].mean()
        st.metric("⭐ Average Score", f"{avg_score:.1f}" if not pd.isna(avg_score) else "N/A")
    
    with col3:
        total_agents = filtered_data['Name of Call Centre Officer'].nunique()
        st.metric("👥 Unique Agents", total_agents)
    
    with col4:
        comments_count = filtered_data['Do you have any other comments you would like to share?'].str.strip().ne('').sum()
        st.metric("💬 Comments Provided", comments_count)
    
    # Tabs for different views
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📈 Overview",
        "👥 Agent Performance",
        "📅 Timeline",
        "📝 Comments",
        "🗂️ Raw Data"
    ])
    
    with tab1:
        col1, col2 = st.columns(2)
        
        with col1:
            fig_dist = px.histogram(
                filtered_data,
                x='Total Score',
                nbins=20,
                title='Total Score Distribution',
                labels={'Total Score': 'Score', 'count': 'Number of Calls'},
                color_discrete_sequence=['#006400']
            )
            fig_dist.update_layout(
                showlegend=False,
                font=dict(family="Arial, sans-serif", size=14),
                title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            )
            st.plotly_chart(fig_dist, use_container_width=True)
        
        with col2:
            avg_score_display = filtered_data['Total Score'].mean()
            if pd.isna(avg_score_display):
                avg_score_display = 0
            fig_gauge = go.Figure(go.Indicator(
                mode="gauge+number",
                value=avg_score_display,
                domain={'x': [0, 1], 'y': [0, 1]},
                title={'text': "Average Total Score"},
                gauge={
                    'axis': {'range': [0, 100]},
                    'bar': {'color': "#006400"},
                    'steps': [
                        {'range': [0, 50], 'color': "#90ee90"},
                        {'range': [51, 75], 'color': "#38F138"},
                        {'range': [76, 100], 'color': "#258925"}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': 144
                    }
                }
            ))
            st.plotly_chart(fig_gauge, use_container_width=True)
    
    with tab2:
        agent_stats = filtered_data.groupby('Name of Call Centre Officer').agg({
            'Total Score': ['mean', 'std', 'count'],
            'Do you have any other comments you would like to share?': lambda x: (x.str.strip() != '').sum()
        }).reset_index()
        agent_stats.columns = ['Agent', 'Avg_Score', 'Std_Dev', 'Call_Count', 'Comments_Count']
        agent_stats = agent_stats.sort_values('Avg_Score', ascending=False)
        
        col1, col2 = st.columns(2)
        
        with col1:
            fig_agent = px.bar(
                agent_stats,
                x='Agent',
                y='Avg_Score',
                title='Average Total Score by Agent',
                labels={'Avg_Score': 'Average Score', 'Agent': 'Agent Name'},
                color='Avg_Score',
                color_continuous_scale='RdYlGn',
                range_color=[0, 100]
            )
            st.plotly_chart(fig_agent, use_container_width=True)
        
        with col2:
            fig_calls = px.bar(
                agent_stats,
                x='Agent',
                y='Call_Count',
                title='Number of Calls Audited by Agent',
                labels={'Call_Count': 'Call Count', 'Agent': 'Agent Name'},
                color='Call_Count',
                color_continuous_scale='Blues'
            )
            st.plotly_chart(fig_calls, use_container_width=True)
        
        st.subheader("Agent Performance Summary")
        st.dataframe(
            agent_stats.rename(columns={
                'Avg_Score': 'Avg Score',
                'Std_Dev': 'Std Dev',
                'Call_Count': 'Calls',
                'Comments_Count': 'Comments'
            }),
            use_container_width=True,
            hide_index=True
        )
    
    with tab3:
        timeline_data = filtered_data.copy()
        if 'Date of interaction' in timeline_data.columns:
            timeline_data['TimelineDate'] = pd.to_datetime(timeline_data['Date of interaction'], errors='coerce').dt.date
        elif 'Audit_Timestamp' in timeline_data.columns:
            timeline_data['TimelineDate'] = pd.to_datetime(timeline_data['Audit_Timestamp'], errors='coerce').dt.date
        else:
            timeline_data['TimelineDate'] = None
        
        daily_stats = timeline_data.groupby('TimelineDate').agg({
            'Total Score': ['mean', 'count']
        }).reset_index()
        daily_stats.columns = ['Date', 'Avg_Score', 'Call_Count']
        
        col1, col2 = st.columns(2)
        
        with col1:
            fig_timeline = px.line(
                daily_stats,
                x='Date',
                y='Avg_Score',
                title='Average Total Score Over Time',
                labels={'Avg_Score': 'Average Score', 'Date': 'Date'},
                markers=True,
                line_shape='linear',
                color_discrete_sequence=['#006400']
            )
            st.plotly_chart(fig_timeline, use_container_width=True)
        
        with col2:
            fig_volume = px.bar(
                daily_stats,
                x='Date',
                y='Call_Count',
                title='Audit Volume Over Time',
                labels={'Call_Count': 'Number of Calls', 'Date': 'Date'},
                color_discrete_sequence=['#006400']
            )
            st.plotly_chart(fig_volume, use_container_width=True)
    
    with tab4:
        comments_col = 'Do you have any other comments you would like to share?'
        base_cols = ['Name of Call Centre Officer', 'Total Score', 'QA Score', comments_col]
        file_col = None
        for candidate in ['File_Name', 'File Name', 'filename', 'File', 'Base']:
            if candidate in filtered_data.columns:
                file_col = candidate
                break
        display_cols = base_cols.copy()
        if file_col:
            display_cols.insert(0, file_col)
        display_cols = [col for col in display_cols if col in filtered_data.columns]
        
        comments_data = filtered_data[
            filtered_data[comments_col].str.strip() != ''
        ][display_cols].copy()
        
        if len(comments_data) > 0:
            st.subheader(f"Comments ({len(comments_data)} entries)")
            
            score_filter = st.select_slider(
                "Filter by Score",
                options=sorted(comments_data['Total Score'].unique())
            )
            comments_filtered = comments_data[comments_data['Total Score'] == score_filter]
            
            for idx, row in comments_filtered.iterrows():
                with st.container():
                    col1, col2, col3 = st.columns([1, 1, 3])
                    with col1:
                        st.write(f"**Score:** {row['Total Score']}/100")
                    with col2:
                        st.write(f"**Agent:** {row['Name of Call Centre Officer']}")
                    with col3:
                        st.write(f"**File:** {row[file_col] if file_col and file_col in row else 'N/A'}")
                    st.info(f"{row[comments_col]}")
                    st.divider()
        else:
            st.info("No comments found for the selected filters.")
    
    with tab5:
        st.subheader("Raw Audit Data")
        
        display_cols = ['Id', 'Name', 'Date of interaction', 'Name of Call Centre Officer',
                        'Name of caller', 'Purpose of call',
                        'Did CCO open the call using the appropriate greetings?',
                        'Was the CCO able to identify and verify the needs of the customer?',
                        'Was the CCO able to educate & inform the customer about the query/enquiry/request',
                        'Did the CCO ensure and confirm the necessary steps to query resolution?',
                        'Was the CCO polite & courteous?', 'Total Score', 'QI Score',
                        'Do you have any other comments you would like to share?', 'Audit_Timestamp']
        
        available_cols = [col for col in display_cols if col in filtered_data.columns]
        
        st.dataframe(
            filtered_data[available_cols],
            use_container_width=True,
            hide_index=True
        )
        
        # Download button - Excel format matching template structure with formulas, colors, and spacing
        excel_buffer = io.BytesIO()
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl import load_workbook
        from datetime import datetime
        
        # Get current month/year for sheet name
        current_date = datetime.now()
        month_year = current_date.strftime("%b %Y")
        
        # Quality columns definition (in order)
        quality_columns = [
            ("Call opening", "Did CCO open the call using the appropriate greetings?"),
            ("Call closure", "Did the CCO end the call politely and professionally?"),
            ("Identification of customer needs", "Was the CCO able to identify and verify the needs of the customer?"),
            ("Educate & Inform", "Was the CCO able to educate & inform the customer about the query/enquiry/request"),
            ("Necessary steps to query resolution", "Did the CCO ensure and confirm the necessary steps to query resolution?"),
            ("Initiative", "Did the CCO accept responsibility for the query? (CAN DO)"),
            ("Identifying further needs", "Did the CCO ask if you had any further needs?"),
            ("Effective communication", "Did the CCO speak clearly and fluently throughout the call?"),
            ("Professionalism", "Was the CCO professional?"),
            ("Effective listening & troubleshooting", "Was the CCO able to communicate effectively through effective listening and troubleshooting?"),
            ("Politeness & courtesy", "Was the CCO polite & courteous?"),
            ("Empathy", "Did the CCO show empathy? (introduce solution statement)")
        ]
        
        # Define colors (RGB hex codes from template)
        green_header = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        light_green_fill = PatternFill(start_color='A9D18E', end_color='A9D18E', fill_type='solid')
        gold_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        white_font = Font(color='FFFFFF', bold=True, size=10)
        black_font = Font(color='000000', size=10)
        bold_black_font = Font(color='000000', bold=True, size=10)
        
        # Define thin border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # ===== SHEET 1: Main Audit Data =====
            export_cols = [col for col in filtered_data.columns if col not in ['Base', 'Contact']]
            filtered_data[export_cols].to_excel(writer, sheet_name=month_year, index=False, startrow=4)
            
            # Format Sheet 1 headers with colors and borders
            ws1 = writer.sheets[month_year]
            for col_idx, col_name in enumerate(export_cols, start=1):
                cell = ws1.cell(row=5, column=col_idx)
                cell.fill = green_header
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Format Sheet 1 data rows with alternating colors
            for row_idx in range(6, len(filtered_data) + 6):
                for col_idx in range(1, len(export_cols) + 1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(export_cols, start=1):
                ws1.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)].width = 18
            
            # Add scoring metrics to Sheet 1 (L2:W3)
            # Row 2: Max scores for each metric
            max_scores = [3, 4, 15, 5, 5, 5, 15, 15, 10, 15, 5, 3]
            for col_idx, score in enumerate(max_scores, start=12):  # Start at column L (12)
                cell = ws1.cell(row=2, column=col_idx, value=score)
                cell.font = bold_black_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = blue_fill
                cell.border = thin_border
            
            # Row 3: Performance percentages for each metric (calculated from AVERAGE/max_score)
            col_letters = ['L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
            for col_idx, col_letter in enumerate(col_letters, start=12):  # Start at column L (12)
                cell = ws1.cell(row=3, column=col_idx)
                cell.value = f'=AVERAGE({col_letter}6:{col_letter}85)/{col_letter}2'
                cell.number_format = '0%'
                cell.font = bold_black_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = light_green_fill
                cell.border = thin_border
            
            # Add column X header (Total Score column)
            header_cell = ws1.cell(row=5, column=24, value='Total Score')
            header_cell.fill = green_header
            header_cell.font = white_font
            header_cell.border = thin_border
            header_cell.alignment = Alignment(horizontal='center')
            
            # Add formula to column X for all data rows (starting from row 6)
            for row_idx in range(6, len(filtered_data) + 6):
                formula_cell = ws1.cell(row=row_idx, column=24)
                formula_cell.value = f'=SUM(IF(L{row_idx}="NA", L$2,IF(L{row_idx}="FATAL", 0, VALUE(L{row_idx}))), IF(M{row_idx}="NA", M$2,IF(M{row_idx}="FATAL", 0, VALUE(M{row_idx}))), IF(N{row_idx}="NA", N$2,IF(N{row_idx}="FATAL", 0, VALUE(N{row_idx}))), IF(O{row_idx}="NA", O$2,IF(O{row_idx}="FATAL", 0, VALUE(O{row_idx}))), IF(P{row_idx}="NA", P$2,IF(P{row_idx}="FATAL", 0, VALUE(P{row_idx}))), IF(Q{row_idx}="NA", Q$2,IF(Q{row_idx}="FATAL", 0, VALUE(Q{row_idx}))), IF(R{row_idx}="NA", R$2,IF(R{row_idx}="FATAL", 0, VALUE(R{row_idx}))), IF(S{row_idx}="NA", S$2,IF(S{row_idx}="FATAL", 0, VALUE(S{row_idx}))), IF(T{row_idx}="NA", T$2,IF(T{row_idx}="FATAL", 0, VALUE(T{row_idx}))), IF(U{row_idx}="NA", U$2,IF(U{row_idx}="FATAL", 0, VALUE(U{row_idx}))),IF(V{row_idx}="NA", V$2,IF(V{row_idx}="FATAL", 0, VALUE(V{row_idx}))), IF(W{row_idx}="NA", W$2,IF(W{row_idx}="FATAL", 0, VALUE(W{row_idx}))))'
                formula_cell.font = black_font
                formula_cell.border = thin_border
                formula_cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    formula_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Add column Y header (Fatal Flag column)
            header_cell_y = ws1.cell(row=5, column=25, value='Fatal Flag')
            header_cell_y.fill = green_header
            header_cell_y.font = white_font
            header_cell_y.border = thin_border
            header_cell_y.alignment = Alignment(horizontal='center')
            
            # Add formula to column Y for all data rows (starting from row 6)
            for row_idx in range(6, len(filtered_data) + 6):
                formula_cell_y = ws1.cell(row=row_idx, column=25)
                formula_cell_y.value = f'=IF(COUNTIF(L{row_idx}:W{row_idx},"Fatal") > 0,"Fatal","-")'
                formula_cell_y.font = black_font
                formula_cell_y.border = thin_border
                formula_cell_y.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    formula_cell_y.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Add column Z header (Score % column)
            header_cell_z = ws1.cell(row=5, column=26, value='Score %')
            header_cell_z.fill = green_header
            header_cell_z.font = white_font
            header_cell_z.border = thin_border
            header_cell_z.alignment = Alignment(horizontal='center')
            
            # Add formula to column Z for all data rows (starting from row 6)
            for row_idx in range(6, len(filtered_data) + 6):
                formula_cell_z = ws1.cell(row=row_idx, column=26)
                formula_cell_z.value = f'=IF(Y{row_idx}="Fatal", "0.00%",X{row_idx}/100)'
                formula_cell_z.font = black_font
                formula_cell_z.border = thin_border
                formula_cell_z.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    formula_cell_z.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # ===== SHEET 2: Individual Performance =====
            if 'QA Score' not in filtered_data.columns:
                filtered_data['QA Score'] = float('nan')
            
            agent_perf = filtered_data.groupby('Name of Call Centre Officer').agg({
                'QA Score': 'mean',
                'Id': 'count'
            }).reset_index()
            agent_perf.columns = ['Row Labels', 'Average of QI Score', 'Count of Name of Call Centre Officer']
            agent_perf = agent_perf.sort_values('Average of QI Score', ascending=False)
            
            grand_total = pd.DataFrame({
                'Row Labels': ['Grand Total'],
                'Average of QI Score': [filtered_data['QA Score'].mean()],
                'Count of Name of Call Centre Officer': [len(filtered_data)]
            })
            agent_perf = pd.concat([agent_perf, grand_total], ignore_index=True)
            
            ws_ip = writer.book.create_sheet('Individual performance')
            # Rows 1-2 blank, headers in row 3, data from row 4
            for col_idx, header in enumerate(agent_perf.columns, start=1):
                cell = ws_ip.cell(row=3, column=col_idx, value=header)
                cell.border = thin_border
                cell.font = bold_black_font
            for row_idx, row in enumerate(agent_perf.values, start=4):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws_ip.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
            
            # ===== SHEET 3: Trend =====
            ws_trend = writer.book.create_sheet('Trend')
            total_records = len(filtered_data)
            
            # Calculate performance for each metric
            metric_performance = []
            for short_name, full_name in quality_columns:
                if full_name in filtered_data.columns:
                    yes_count = (filtered_data[full_name] == 'Yes').sum()
                else:
                    yes_count = 0
                performance = (yes_count / total_records) if total_records > 0 else 0
                metric_performance.append((short_name, performance))
            
            # Row 2: Headers with formatting and borders
            b2_cell = ws_trend.cell(row=2, column=2, value='Call Centre Productivity')
            b2_cell.fill = green_header
            b2_cell.font = white_font
            b2_cell.alignment = Alignment(horizontal='center')
            b2_cell.border = thin_border
            
            c2_cell = ws_trend.cell(row=2, column=3, value=current_date.date())
            c2_cell.fill = green_header
            c2_cell.font = white_font
            c2_cell.alignment = Alignment(horizontal='center')
            c2_cell.border = thin_border
            
            # Metric names starting at F2 with appropriate colors and borders
            metric_colors = [
                (0, blue_fill), (1, blue_fill),  # Call opening, Call closure - blue
                (2, light_green_fill), (3, light_green_fill), (4, light_green_fill), (5, light_green_fill), (6, light_green_fill),  # Problem solving - light green
                (7, gold_fill), (8, gold_fill), (9, gold_fill), (10, gold_fill), (11, gold_fill)  # Soft skills - gold
            ]
            for idx, (short_name, _) in enumerate(metric_performance):
                cell = ws_trend.cell(row=2, column=6+idx, value=short_name)
                cell.fill = metric_colors[idx][1]
                cell.font = bold_black_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Row 3: Summary data with borders
            cell_b3 = ws_trend.cell(row=3, column=2, value='QI Audit')
            cell_b3.font = black_font
            cell_b3.alignment = Alignment(horizontal='left')
            cell_b3.border = thin_border
            
            cell_c3 = ws_trend.cell(row=3, column=3, value=total_records)
            cell_c3.font = black_font
            cell_c3.alignment = Alignment(horizontal='right')
            cell_c3.border = thin_border
            
            ws_trend.cell(row=3, column=5, value='Overall performance').font = black_font
            
            # Performance values starting at F3 with borders
            for idx, (_, perf) in enumerate(metric_performance):
                cell = ws_trend.cell(row=3, column=6+idx, value=perf)
                cell.font = black_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Row 4: Fatal count with borders
            cell_b4 = ws_trend.cell(row=4, column=2, value='Fatal count')
            cell_b4.font = black_font
            cell_b4.border = thin_border
            
            fatal_count = 0
            for col in filtered_data.columns:
                try:
                    fatal_count += (filtered_data[col] == 'Fatal').sum()
                except:
                    pass
            cell_c4 = ws_trend.cell(row=4, column=3, value=int(fatal_count))
            cell_c4.font = black_font
            cell_c4.alignment = Alignment(horizontal='right')
            cell_c4.border = thin_border
            
            # Row 5: QI Score with borders
            cell_b5 = ws_trend.cell(row=5, column=2, value='QI Score')
            cell_b5.font = black_font
            cell_b5.border = thin_border
            
            avg_qi = filtered_data['QA Score'].mean() if len(filtered_data) > 0 else 0
            cell_c5 = ws_trend.cell(row=5, column=3, value=round(avg_qi, 2))
            cell_c5.font = black_font
            cell_c5.alignment = Alignment(horizontal='right')
            cell_c5.border = thin_border
            
            # ===== SHEET 4: Team Performance =====
            ws_team = writer.book.create_sheet('Team performance')
            
            # Row 2: Category headers
            c2 = ws_team.cell(row=2, column=3, value='Overall performance')
            c2.font = bold_black_font
            g2 = ws_team.cell(row=2, column=7, value='Introduction & Conclusion')
            g2.font = bold_black_font
            j2 = ws_team.cell(row=2, column=10, value='Problem solving')
            j2.font = bold_black_font
            n2 = ws_team.cell(row=2, column=14, value='Soft skills')
            n2.font = bold_black_font
            
            # Data rows with color coding and borders
            all_metrics_sorted = sorted(
                [(name, perf) for name, perf in metric_performance],
                key=lambda x: x[1],
                reverse=True
            )
            for row_idx, (metric_name, perf) in enumerate(all_metrics_sorted, start=3):
                # Column B: Metric name with border
                cell_b = ws_team.cell(row=row_idx, column=2, value=metric_name)
                cell_b.font = black_font
                cell_b.border = thin_border
                # Determine color based on metric category
                if metric_name in ['Call opening', 'Call closure']:
                    cell_b.fill = blue_fill
                elif metric_name in ['Identification of customer needs', 'Educate & Inform', 'Necessary steps to query resolution', 'Initiative', 'Identifying further needs']:
                    cell_b.fill = light_green_fill
                else:
                    cell_b.fill = gold_fill
                
                # Column C: Performance value with border
                cell_c = ws_team.cell(row=row_idx, column=3, value=perf)
                cell_c.font = black_font
                cell_c.border = thin_border
            
            # Introduction & Conclusion (F3:G) with borders
            intro_data = [
                ('Call opening', metric_performance[0][1]),
                ('Call closure', metric_performance[1][1])
            ]
            for row_idx, (metric_name, perf) in enumerate(intro_data, start=3):
                cell_f = ws_team.cell(row=row_idx, column=6, value=metric_name)
                cell_f.fill = blue_fill
                cell_f.font = black_font
                cell_f.border = thin_border
                cell_g = ws_team.cell(row=row_idx, column=7, value=perf)
                cell_g.font = black_font
                cell_g.border = thin_border
            
            # Problem solving (I3:J) with borders
            problem_data = [
                ('Identification of customer needs', metric_performance[2][1]),
                ('Educate & Inform', metric_performance[3][1]),
                ('Necessary steps to query resolution', metric_performance[4][1]),
                ('Initiative', metric_performance[5][1]),
                ('Identifying further needs', metric_performance[6][1])
            ]
            for row_idx, (metric_name, perf) in enumerate(problem_data, start=3):
                cell_i = ws_team.cell(row=row_idx, column=9, value=metric_name)
                cell_i.fill = light_green_fill
                cell_i.font = black_font
                cell_i.border = thin_border
                cell_j = ws_team.cell(row=row_idx, column=10, value=perf)
                cell_j.font = black_font
                cell_j.border = thin_border
            
            # Soft skills (M3:N) with borders
            soft_skills_data = [
                ('Effective communication', metric_performance[7][1]),
                ('Professionalism', metric_performance[8][1]),
                ('Effective listening & troubleshooting', metric_performance[9][1]),
                ('Politeness & courtesy', metric_performance[10][1]),
                ('Empathy', metric_performance[11][1])
            ]
            for row_idx, (metric_name, perf) in enumerate(soft_skills_data, start=3):
                cell_m = ws_team.cell(row=row_idx, column=13, value=metric_name)
                cell_m.fill = gold_fill
                cell_m.font = black_font
                cell_m.border = thin_border
                cell_n = ws_team.cell(row=row_idx, column=14, value=perf)
                cell_n.font = black_font
                cell_n.border = thin_border
            
            # ===== SHEET 5: Charts & Summary =====
            from openpyxl.chart import BarChart, LineChart, Reference, PieChart
            ws_charts = writer.book.create_sheet('Charts & Summary')
            
            # Organize metrics by category
            intro_conclusion = [
                ('Call opening', metric_performance[0][1]),
                ('Call closure', metric_performance[1][1])
            ]
            
            problem_solving = [
                ('Identification of customer needs', metric_performance[2][1]),
                ('Educate & Inform', metric_performance[3][1]),
                ('Necessary steps to query resolution', metric_performance[4][1]),
                ('Initiative', metric_performance[5][1]),
                ('Identifying further needs', metric_performance[6][1])
            ]
            
            soft_skills = [
                ('Effective communication', metric_performance[7][1]),
                ('Professionalism', metric_performance[8][1]),
                ('Effective listening & troubleshooting', metric_performance[9][1]),
                ('Politeness & courtesy', metric_performance[10][1]),
                ('Empathy', metric_performance[11][1])
            ]
            
            # ===== Chart 1: Overall Performance =====
            chart1_row = 2
            ws_charts.cell(row=chart1_row, column=1, value='Overall Performance').font = Font(bold=True, size=11)
            for idx, (metric_name, perf) in enumerate(metric_performance, start=chart1_row + 1):
                ws_charts.cell(row=idx, column=1, value=metric_name).font = black_font
                ws_charts.cell(row=idx, column=1).border = thin_border
                score_cell = ws_charts.cell(row=idx, column=2, value=perf)
                score_cell.font = black_font
                score_cell.number_format = '0%'
                score_cell.border = thin_border
                if perf >= 0.75:
                    score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif perf >= 0.50:
                    score_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                else:
                    score_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # Create Overall Performance chart
            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = "Overall Performance"
            chart1.y_axis.title = 'Performance %'
            data1 = Reference(ws_charts, min_col=2, min_row=chart1_row, max_row=len(metric_performance) + chart1_row)
            cats1 = Reference(ws_charts, min_col=1, min_row=chart1_row + 1, max_row=len(metric_performance) + chart1_row)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)
            chart1.height = 9
            chart1.width = 18
            chart1.series[0].graphicalProperties.solidFill = "00B0F0"
            ws_charts.add_chart(chart1, "D2")
            
            # ===== Chart 2: Introduction and Conclusion =====
            chart2_row = chart1_row + len(metric_performance) + 3
            ws_charts.cell(row=chart2_row, column=1, value='Introduction & Conclusion').font = Font(bold=True, size=11)
            for idx, (metric_name, perf) in enumerate(intro_conclusion, start=chart2_row + 1):
                ws_charts.cell(row=idx, column=1, value=metric_name).font = black_font
                ws_charts.cell(row=idx, column=1).fill = blue_fill
                ws_charts.cell(row=idx, column=1).border = thin_border
                score_cell = ws_charts.cell(row=idx, column=2, value=perf)
                score_cell.font = black_font
                score_cell.number_format = '0%'
                score_cell.border = thin_border
                if perf >= 0.75:
                    score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif perf >= 0.50:
                    score_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                else:
                    score_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # Create Introduction & Conclusion chart
            chart2 = BarChart()
            chart2.type = "col"
            chart2.title = "Introduction & Conclusion"
            chart2.y_axis.title = 'Performance %'
            data2 = Reference(ws_charts, min_col=2, min_row=chart2_row, max_row=chart2_row + len(intro_conclusion))
            cats2 = Reference(ws_charts, min_col=1, min_row=chart2_row + 1, max_row=chart2_row + len(intro_conclusion))
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            chart2.height = 9
            chart2.width = 18
            chart2.series[0].graphicalProperties.solidFill = "00B0F0"
            ws_charts.add_chart(chart2, "J2")
            
            # ===== Chart 3: Problem Solving =====
            chart3_row = chart2_row + len(intro_conclusion) + 3
            ws_charts.cell(row=chart3_row, column=1, value='Problem Solving').font = Font(bold=True, size=11)
            for idx, (metric_name, perf) in enumerate(problem_solving, start=chart3_row + 1):
                ws_charts.cell(row=idx, column=1, value=metric_name).font = black_font
                ws_charts.cell(row=idx, column=1).fill = light_green_fill
                ws_charts.cell(row=idx, column=1).border = thin_border
                score_cell = ws_charts.cell(row=idx, column=2, value=perf)
                score_cell.font = black_font
                score_cell.number_format = '0%'
                score_cell.border = thin_border
                if perf >= 0.75:
                    score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif perf >= 0.50:
                    score_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                else:
                    score_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # Create Problem Solving chart
            chart3 = BarChart()
            chart3.type = "col"
            chart3.title = "Problem Solving"
            chart3.y_axis.title = 'Performance %'
            data3 = Reference(ws_charts, min_col=2, min_row=chart3_row, max_row=chart3_row + len(problem_solving))
            cats3 = Reference(ws_charts, min_col=1, min_row=chart3_row + 1, max_row=chart3_row + len(problem_solving))
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(cats3)
            chart3.height = 9
            chart3.width = 18
            chart3.series[0].graphicalProperties.solidFill = "A9D18E"
            ws_charts.add_chart(chart3, "D" + str(chart2_row + 2))
            
            # ===== Chart 4: Soft Skills =====
            chart4_row = chart3_row + len(problem_solving) + 3
            ws_charts.cell(row=chart4_row, column=1, value='Soft Skills').font = Font(bold=True, size=11)
            for idx, (metric_name, perf) in enumerate(soft_skills, start=chart4_row + 1):
                ws_charts.cell(row=idx, column=1, value=metric_name).font = black_font
                ws_charts.cell(row=idx, column=1).fill = gold_fill
                ws_charts.cell(row=idx, column=1).border = thin_border
                score_cell = ws_charts.cell(row=idx, column=2, value=perf)
                score_cell.font = black_font
                score_cell.number_format = '0%'
                score_cell.border = thin_border
                if perf >= 0.75:
                    score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif perf >= 0.50:
                    score_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                else:
                    score_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # Create Soft Skills chart
            chart4 = BarChart()
            chart4.type = "col"
            chart4.title = "Soft Skills"
            chart4.y_axis.title = 'Performance %'
            data4 = Reference(ws_charts, min_col=2, min_row=chart4_row, max_row=chart4_row + len(soft_skills))
            cats4 = Reference(ws_charts, min_col=1, min_row=chart4_row + 1, max_row=chart4_row + len(soft_skills))
            chart4.add_data(data4, titles_from_data=True)
            chart4.set_categories(cats4)
            chart4.height = 9
            chart4.width = 18
            chart4.series[0].graphicalProperties.solidFill = "FFC000"
            ws_charts.add_chart(chart4, "J" + str(chart2_row + 2))
            
            # Set column widths
            ws_charts.column_dimensions['A'].width = 40
            ws_charts.column_dimensions['B'].width = 15
        
        excel_buffer.seek(0)

        st.download_button(
            label="📊 Download Report as Excel",
            data=excel_buffer,
            file_name=f"qa_audit_report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
