from openpyxl import load_workbook
from openpyxl.styles import Border

wb = load_workbook('Score Sheet/Call Centre Audit Report - October 2024 (Recovered).xlsx')

print("=" * 100)
print("BORDER ANALYSIS FROM TEMPLATE")
print("=" * 100)

# Check Trend sheet for borders
print("\n*** TREND SHEET - BORDERS ***")
ws_trend = wb['Trend']
for row in range(1, 7):
    for col in range(1, 19):
        cell = ws_trend.cell(row=row, column=col)
        if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
            print(f"{cell.coordinate}: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")

# Check Team Performance sheet for borders
print("\n*** TEAM PERFORMANCE SHEET - BORDERS ***")
ws_team = wb['Team performance']
for row in range(1, 13):
    for col in range(1, 15):
        cell = ws_team.cell(row=row, column=col)
        if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
            print(f"{cell.coordinate}: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")

# Check Individual Performance sheet for borders
print("\n*** INDIVIDUAL PERFORMANCE SHEET - BORDERS ***")
ws_ip = wb['Individual performance']
for row in range(1, 10):
    for col in range(1, 4):
        cell = ws_ip.cell(row=row, column=col)
        if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
            print(f"{cell.coordinate}: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")

wb.close()
