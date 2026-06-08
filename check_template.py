from openpyxl import load_workbook

wb = load_workbook('Score Sheet/Call Centre Audit Report - October 2024 (Recovered).xlsx')
ws = wb['Oct 2024']

print("OCT 2024 SHEET - First 5 rows, all columns A-AA:")
for row in range(1, 6):
    row_data = []
    for col in range(1, 28):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val:
            if isinstance(val, float):
                row_data.append(f"{val:.2f}")
            else:
                row_data.append(str(val)[:15])
        else:
            row_data.append("")
    print(f"Row {row}: {row_data}")

print("\nTotal rows:", ws.max_row)
print("Total columns:", ws.max_column)

# Show individual performance
print("\n\nINDIVIDUAL PERFORMANCE SHEET:")
ws_ip = wb['Individual performance']
for row in range(1, 10):
    row_data = []
    for col in range(1, 4):
        cell = ws_ip.cell(row=row, column=col)
        val = cell.value
        if val:
            if isinstance(val, float):
                row_data.append(f"{val:.4f}")
            else:
                row_data.append(str(val)[:30])
        else:
            row_data.append("")
    print(f"Row {row}: {row_data}")

wb.close()
