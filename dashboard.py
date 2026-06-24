import streamlit as st
import os
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import glob
import io
import re
from extra_streamlit_components import CookieManager

st.set_page_config(page_title="QA Audit Dashboard", layout="wide")

# --- User Authentication and RBAC ---
import streamlit_authenticator as stauth

# Define credentials dictionary for streamlit-authenticator >=0.4.2
credentials = {
    "usernames": {
        "amanda.bio-marfo@nationwidemh.com": {
            "name": "Amanda Bio-Marfo",
            "password": stauth.Hasher.hash("123")
        },
        "suraiyatu.mohammed@nationwidemh.com": {
            "name": "Surayatu Mohammed",
            "password": stauth.Hasher.hash("456")
        },
        "edem.dzitrie@nationwidemh.com": {
            "name": "Edem Dzitrie",
            "password": stauth.Hasher.hash("789")
        }
    }
}

cookie_manager = CookieManager(key="qa_app_dashboard_cookie_manager")

authenticator = stauth.Authenticate(
    credentials=credentials,
    cookie_name="qa_app_dashboard",
    key="qa_app_dashboard_auth",
    expiry_days=1,
    cookie_manager=cookie_manager
)

# Login widget
authenticator.login(location='main')

# Get values from session state
name = st.session_state.get("name")
authentication_status = st.session_state.get("authentication_status")
username = st.session_state.get("username")

def get_user_role(username):
    # Example static mapping; replace with DB lookup as needed
    role_map = {
        "amanda.bio-marfo@nationwidemh.com": "admin",
        "suraiyatu.mohammed@nationwidemh.com": "auditor",
        "edem.dzitrie@nationwidemh.com": "supervisor"
    }
    return role_map.get(username, None)

def show_admin_panel():
    st.info("Admin Panel: You have full access.")

def show_approval_dashboard():
    st.info("Supervisor Dashboard: You have full access.")

def show_audit_form():
    st.info("Auditor Form: You have full access.")

if authentication_status == False:
    st.error("Invalid credentials")
    st.stop()
elif authentication_status is None:
    st.warning("Please enter your username and password")
    st.stop()

# If we reach here, user is authenticated
st.success(f"Welcome {name}")
user_role = get_user_role(username)
if user_role == "admin":
    show_admin_panel()
elif user_role == "supervisor":
    show_approval_dashboard()
elif user_role == "auditor":
    show_audit_form()

# Load and combine all audit log files into audit_data
audit_files = glob.glob("Audit_log_calls/audit_log_*.xlsx")
if not audit_files:
    st.error("No audit log files found. Please ensure audit_log_*.xlsx files are present.")
    st.stop()
dfs = []
for file in audit_files:
    try:
        df = pd.read_excel(file)
        dfs.append(df)
    except Exception as e:
        st.warning(f"Could not read {file}: {e}")
if not dfs:
    st.error("No valid audit log data loaded.")
    st.stop()
audit_data = pd.concat(dfs, ignore_index=True)

st.title("📊 QA Audit Dashboard")

# Date range filter setup (must come after audit_data is defined)

# Try to infer a consistent datetime format for 'Audit_Timestamp' to avoid the warning
if 'Audit_Timestamp' in audit_data.columns:
    ts = pd.to_datetime(audit_data['Audit_Timestamp'], format="%d/%m/%Y %H:%M", errors='coerce')
elif 'Date of interaction' in audit_data.columns:
    ts = pd.to_datetime(audit_data['Date of interaction'], format="%d/%m/%Y %H:%M", errors='coerce')
else:
    ts = None

date_range = None
if ts is not None and ts.notna().any():
    min_dt = ts.min().date()
    max_dt = ts.max().date()
    date_range = st.sidebar.date_input(
        "📅 Date Range",
        value=(min_dt, max_dt),
        min_value=min_dt,
        max_value=max_dt
    )


    # Display logo in the top right corner
    logo_path = "logo.png"
    col_title, col_logo = st.columns([6, 1])


    with col_logo:
        st.image(logo_path, use_column_width=True)
filtered_data = audit_data.copy()
if ts is not None and isinstance(date_range, (list, tuple)) and len(date_range) == 2:
    start_date, end_date = date_range
    if start_date is not None and end_date is not None:
        valid_ts = ts.notna()
        start_ts = pd.to_datetime(start_date)
        end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        filtered_data = filtered_data[valid_ts & (ts >= start_ts) & (ts <= end_ts)]
# ...existing code...


# Standardize agent column for consistency
if 'Agent' not in filtered_data.columns and 'Name of Call Centre Officer' in filtered_data.columns:
    filtered_data['Agent'] = filtered_data['Name of Call Centre Officer']
elif 'Agent' in filtered_data.columns and 'Name of Call Centre Officer' not in filtered_data.columns:
    filtered_data['Name of Call Centre Officer'] = filtered_data['Agent']

# Agent filter
agents = sorted(filtered_data['Agent'].dropna().unique())
selected_agents = st.sidebar.multiselect("👥 Select Agents", agents, default=agents)
if selected_agents:
    filtered_data = filtered_data[filtered_data['Agent'].isin(selected_agents)]

# Ensure Total Score is numeric and comments are strings
filtered_data = filtered_data.copy()
if 'Total Score' in filtered_data.columns:
    filtered_data['Total Score'] = pd.to_numeric(filtered_data['Total Score'], errors='coerce')
if 'QA Score' in filtered_data.columns:
    filtered_data['QA Score'] = pd.to_numeric(filtered_data['QA Score'], errors='coerce')
if 'Do you have any other comments you would like to share?' in filtered_data.columns:
    filtered_data['Do you have any other comments you would like to share?'] = filtered_data['Do you have any other comments you would like to share?'].fillna('').astype(str)
else:
    filtered_data['Do you have any other comments you would like to share?'] = ''


# Score threshold (dynamic min/max to avoid RangeError)
if 'Total Score' in filtered_data.columns and filtered_data['Total Score'].notna().any():
    min_score = int(filtered_data['Total Score'].min())
    max_score = int(filtered_data['Total Score'].max())
    if min_score == max_score:
        # If all scores are the same, set a small valid range
        score_range = st.sidebar.slider("⭐ Total Score Range", min_score, min_score+1, (min_score, min_score+1))
    else:
        score_range = st.sidebar.slider("⭐ Total Score Range", min_score, max_score, (min_score, max_score))
    filtered_data = filtered_data[
        filtered_data['Total Score'].notna() &
        (filtered_data['Total Score'] >= score_range[0]) &
        (filtered_data['Total Score'] <= score_range[1])
    ]
else:
    st.sidebar.info("No scores available for range filter.")

st.sidebar.write(f"📌 Showing {len(filtered_data)} records")

# Main metrics
col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric("📞 Total Calls Audited", len(filtered_data))

with col2:
    avg_score = filtered_data['Total Score'].mean()
    st.metric("⭐ Average Score", f"{avg_score:.1f}" if not pd.isna(avg_score) else "N/A")

with col3:
    total_agents = filtered_data['Name of Call Centre Officer'].nunique()
    st.metric("👥 Unique Agents", total_agents)

with col4:
    comments_count = filtered_data['Do you have any other comments you would like to share?'].str.strip().ne('').sum()
    st.metric("💬 Comments Provided", comments_count)

# Tabs for different views
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📈 Overview",
    "👥 Agent Performance",
    "📅 Timeline",
    "📝 Comments",
    "🗂️ Raw Data"
])

with tab1:
    col1, col2 = st.columns(2)

    with col1:
        fig_dist = px.histogram(
            filtered_data,
            x='Total Score',
            nbins=20,
            title='Total Score Distribution',
            labels={'Total Score': 'Score', 'count': 'Number of Calls'},
            color_discrete_sequence=['#006400']  # dark green
        )
        fig_dist.update_layout(
            showlegend=False,
            font=dict(family="Arial, sans-serif", size=14),
            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
        )
        st.plotly_chart(fig_dist, use_container_width=True)

    with col2:
        # Score gauge
        avg_score_display = filtered_data['Total Score'].mean()
        if pd.isna(avg_score_display):
            avg_score_display = 0
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number",
            value=avg_score_display,
            domain={'x': [0, 1], 'y': [0, 1]},
            title={'text': "Average Total Score"},
            gauge={
                'axis': {'range': [0, 100]},
                'bar': {'color': "#006400"},  # dark green
                'steps': [
                    {'range': [0, 50], 'color': "#90ee90"},  # light green
                    {'range': [51, 75], 'color': "#38F138"}, # forest green
                    {'range': [76, 100], 'color': "#258925"} # dark green
                ],
                'threshold': {
                    'line': {'color': "red", 'width': 4},
                    'thickness': 0.75,
                    'value': 144
                }
            }
        ))
        st.plotly_chart(fig_gauge, use_container_width=True)

    # ...removed Quality Metrics Trend (Grouped) table from Overview tab...

# Tab 2: Agent Performance
with tab2:
    agent_stats = filtered_data.groupby('Name of Call Centre Officer').agg({
        'Total Score': ['mean', 'std', 'count'],
        'Do you have any other comments you would like to share?': lambda x: (x.str.strip() != '').sum()
    }).reset_index()
    agent_stats.columns = ['Agent', 'Avg_Score', 'Std_Dev', 'Call_Count', 'Comments_Count']
    agent_stats = agent_stats.sort_values('Avg_Score', ascending=False)

    col1, col2 = st.columns(2)

    with col1:
        # Agent average scores
        fig_agent = px.bar(
            agent_stats,
            x='Agent',
            y='Avg_Score',
            title='Average Total Score by Agent',
            labels={'Avg_Score': 'Average Score', 'Agent': 'Agent Name'},
            color='Avg_Score',
            color_continuous_scale='RdYlGn',
            range_color=[0, 100]
        )
        fig_agent.update_layout(
            font=dict(family="Arial, sans-serif", size=14),
            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
        )
        st.plotly_chart(fig_agent, use_container_width=True)

    with col2:
        # Call count by agent
        fig_calls = px.bar(
            agent_stats,
            x='Agent',
            y='Call_Count',
            title='Number of Calls Audited by Agent',
            labels={'Call_Count': 'Call Count', 'Agent': 'Agent Name'},
            color='Call_Count',
            color_continuous_scale='Blues'
        )
        fig_calls.update_layout(
            font=dict(family="Arial, sans-serif", size=14),
            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
        )
        st.plotly_chart(fig_calls, use_container_width=True)

    # Agent performance table
    st.subheader("Agent Performance Summary")
    st.dataframe(
        agent_stats.rename(columns={
            'Avg_Score': 'Avg Score',
            'Std_Dev': 'Std Dev',
            'Call_Count': 'Calls',
            'Comments_Count': 'Comments'
        }),
        use_container_width=True,
        hide_index=True
    )

# Tab 3: Timeline
with tab3:
    # Use call date if available, else fallback to audit timestamp
    timeline_data = filtered_data.copy()
    if 'Date of interaction' in timeline_data.columns:
        timeline_data['TimelineDate'] = pd.to_datetime(timeline_data['Date of interaction'], errors='coerce').dt.date
    elif 'Audit_Timestamp' in timeline_data.columns:
        timeline_data['TimelineDate'] = pd.to_datetime(timeline_data['Audit_Timestamp'], errors='coerce').dt.date
    else:
        timeline_data['TimelineDate'] = None

    daily_stats = timeline_data.groupby('TimelineDate').agg({
        'Total Score': ['mean', 'count']
    }).reset_index()
    daily_stats.columns = ['Date', 'Avg_Score', 'Call_Count']

    col1, col2 = st.columns(2)

    with col1:
        fig_timeline = px.line(
            daily_stats,
            x='Date',
            y='Avg_Score',
            title='Average Total Score Over Time',
            labels={'Avg_Score': 'Average Score', 'Date': 'Date'},
            markers=True,
            line_shape='linear',
            color_discrete_sequence=['#006400']  # dark green
        )
        fig_timeline.update_layout(
            font=dict(family="Arial, sans-serif", size=14),
            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
        )
        st.plotly_chart(fig_timeline, use_container_width=True)

    with col2:
        fig_volume = px.bar(
            daily_stats,
            x='Date',
            y='Call_Count',
            title='Audit Volume Over Time',
            labels={'Call_Count': 'Number of Calls', 'Date': 'Date'},
            color_discrete_sequence=['#006400']  # dark green
        )
        fig_volume.update_layout(
            font=dict(family="Arial, sans-serif", size=14),
            title_font=dict(family="Arial, sans-serif", size=18, color="#222"),
            xaxis_title_font=dict(family="Arial, sans-serif", size=16),
            yaxis_title_font=dict(family="Arial, sans-serif", size=16)
        )
        st.plotly_chart(fig_volume, use_container_width=True)

# Tab 4: Comments
with tab4:
    comments_col = 'Do you have any other comments you would like to share?'
    # Robustly select columns for comments display
    base_cols = ['Name of Call Centre Officer', 'Total Score', 'QA Score', comments_col]
    # Try to include File_Name or fallback to another column if present
    file_col = None
    for candidate in ['File_Name', 'File Name', 'filename', 'File', 'Base']:
        if candidate in filtered_data.columns:
            file_col = candidate
            break
    display_cols = base_cols.copy()
    if file_col:
        display_cols.insert(0, file_col)
    # Only keep columns that exist
    display_cols = [col for col in display_cols if col in filtered_data.columns]
    comments_data = filtered_data[
        filtered_data[comments_col].str.strip() != ''
    ][display_cols].copy()
    if len(comments_data) > 0:
        st.subheader(f"Comments ({len(comments_data)} entries)")

        # Filter by score
        score_filter = st.select_slider(
            "Filter by Score",
            options=sorted(comments_data['Total Score'].unique())
        )
        comments_filtered = comments_data[comments_data['Total Score'] == score_filter]

        for idx, row in comments_filtered.iterrows():
            with st.container():
                col1, col2, col3 = st.columns([1, 1, 3])
                with col1:
                    st.write(f"**Score:** {row['Total Score']}/100")
                with col2:
                    st.write(f"**Agent:** {row['Name of Call Centre Officer']}")
                with col3:
                    st.write(f"**File:** {row[file_col] if file_col and file_col in row else 'N/A'}")
                st.info(f"{row[comments_col]}")
                st.divider()
    else:
        st.info("No comments found for the selected filters.")

# Tab 5: Raw Data
with tab5:
    st.subheader("Raw Audit Data")

    # Display key columns
    display_cols = ['Id', 'Name', 'Date of interaction', 'Name of Call Centre Officer',
                    'Name of caller', 'Purpose of call',
                    'Did CCO open the call using the appropriate greetings?',
                    'Was the CCO able to identify and verify the needs of the customer?',
                    'Was the CCO able to educate & inform the customer about the query/enquiry/request',
                    'Did the CCO ensure and confirm the necessary steps to query resolution?',
                    'Was the CCO polite & courteous?', 'Total Score', 'QI Score',
                    'Do you have any other comments you would like to share?', 'Audit_Timestamp']

    # Only show columns that exist in the data
    available_cols = [col for col in display_cols if col in filtered_data.columns]

    st.dataframe(
        filtered_data[available_cols],
        use_container_width=True,
        hide_index=True
    )

    # Download button - Excel format matching template structure with formulas, colors, and spacing
    excel_buffer = io.BytesIO()
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl import load_workbook
    from datetime import datetime
    
    # Get current month/year for sheet name
    current_date = datetime.now()
    month_year = current_date.strftime("%b %Y")
    
    # Quality columns definition (in order)
    quality_columns = [
        ("Call opening", "Did CCO open the call using the appropriate greetings?"),
        ("Call closure", "Did the CCO end the call politely and professionally?"),
        ("Identification of customer needs", "Was the CCO able to identify and verify the needs of the customer?"),
        ("Educate & Inform", "Was the CCO able to educate & inform the customer about the query/enquiry/request"),
        ("Necessary steps to query resolution", "Did the CCO ensure and confirm the necessary steps to query resolution?"),
        ("Initiative", "Did the CCO accept responsibility for the query? (CAN DO)"),
        ("Identifying further needs", "Did the CCO ask if you had any further needs?"),
        ("Effective communication", "Did the CCO speak clearly and fluently throughout the call?"),
        ("Professionalism", "Was the CCO professional?"),
        ("Effective listening & troubleshooting", "Was the CCO able to communicate effectively through effective listening and troubleshooting?"),
        ("Politeness & courtesy", "Was the CCO polite & courteous?"),
        ("Empathy", "Did the CCO show empathy? (introduce solution statement)")
    ]
    
    # Define colors (RGB hex codes from template)
    green_header = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    light_green_fill = PatternFill(start_color='A9D18E', end_color='A9D18E', fill_type='solid')
    gold_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    white_font = Font(color='FFFFFF', bold=True, size=10)
    black_font = Font(color='000000', size=10)
    bold_black_font = Font(color='000000', bold=True, size=10)
    
    # Define thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # ===== SHEET 1: Main Audit Data =====
        export_cols = [col for col in filtered_data.columns if col not in ['Base', 'Contact']]
        filtered_data[export_cols].to_excel(writer, sheet_name=month_year, index=False, startrow=4)
        
        # ===== SHEET 2: Individual Performance =====
        if 'QA Score' not in filtered_data.columns:
            filtered_data['QA Score'] = float('nan')
        
        agent_perf = filtered_data.groupby('Name of Call Centre Officer').agg({
            'QA Score': 'mean',
            'Id': 'count'
        }).reset_index()
        agent_perf.columns = ['Row Labels', 'Average of QI Score', 'Count of Name of Call Centre Officer']
        agent_perf = agent_perf.sort_values('Average of QI Score', ascending=False)
        
        grand_total = pd.DataFrame({
            'Row Labels': ['Grand Total'],
            'Average of QI Score': [filtered_data['QA Score'].mean()],
            'Count of Name of Call Centre Officer': [len(filtered_data)]
        })
        agent_perf = pd.concat([agent_perf, grand_total], ignore_index=True)
        
        ws_ip = writer.book.create_sheet('Individual performance')
        # Rows 1-2 blank, headers in row 3, data from row 4
        for col_idx, header in enumerate(agent_perf.columns, start=1):
            cell = ws_ip.cell(row=3, column=col_idx, value=header)
            cell.border = thin_border
            cell.font = bold_black_font
        for row_idx, row in enumerate(agent_perf.values, start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = ws_ip.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # ===== SHEET 3: Trend =====
        ws_trend = writer.book.create_sheet('Trend')
        total_records = len(filtered_data)
        
        # Calculate performance for each metric
        metric_performance = []
        for short_name, full_name in quality_columns:
            if full_name in filtered_data.columns:
                yes_count = (filtered_data[full_name] == 'Yes').sum()
            else:
                yes_count = 0
            performance = (yes_count / total_records) if total_records > 0 else 0
            metric_performance.append((short_name, performance))
        
        # Row 2: Headers with formatting and borders
        b2_cell = ws_trend.cell(row=2, column=2, value='Call Centre Productivity')
        b2_cell.fill = green_header
        b2_cell.font = white_font
        b2_cell.alignment = Alignment(horizontal='center')
        b2_cell.border = thin_border
        
        c2_cell = ws_trend.cell(row=2, column=3, value=current_date.date())
        c2_cell.fill = green_header
        c2_cell.font = white_font
        c2_cell.alignment = Alignment(horizontal='center')
        c2_cell.border = thin_border
        
        # Metric names starting at F2 with appropriate colors and borders
        metric_colors = [
            (0, blue_fill), (1, blue_fill),  # Call opening, Call closure - blue
            (2, light_green_fill), (3, light_green_fill), (4, light_green_fill), (5, light_green_fill), (6, light_green_fill),  # Problem solving - light green
            (7, gold_fill), (8, gold_fill), (9, gold_fill), (10, gold_fill), (11, gold_fill)  # Soft skills - gold
        ]
        for idx, (short_name, _) in enumerate(metric_performance):
            cell = ws_trend.cell(row=2, column=6+idx, value=short_name)
            cell.fill = metric_colors[idx][1]
            cell.font = bold_black_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Row 3: Summary data with borders
        cell_b3 = ws_trend.cell(row=3, column=2, value='QI Audit')
        cell_b3.font = black_font
        cell_b3.alignment = Alignment(horizontal='left')
        cell_b3.border = thin_border
        
        cell_c3 = ws_trend.cell(row=3, column=3, value=total_records)
        cell_c3.font = black_font
        cell_c3.alignment = Alignment(horizontal='right')
        cell_c3.border = thin_border
        
        ws_trend.cell(row=3, column=5, value='Overall performance').font = black_font
        
        # Performance values starting at F3 with borders
        for idx, (_, perf) in enumerate(metric_performance):
            cell = ws_trend.cell(row=3, column=6+idx, value=perf)
            cell.font = black_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Row 4: Fatal count with borders
        cell_b4 = ws_trend.cell(row=4, column=2, value='Fatal count')
        cell_b4.font = black_font
        cell_b4.border = thin_border
        
        fatal_count = 0
        for col in filtered_data.columns:
            try:
                fatal_count += (filtered_data[col] == 'Fatal').sum()
            except:
                pass
        cell_c4 = ws_trend.cell(row=4, column=3, value=int(fatal_count))
        cell_c4.font = black_font
        cell_c4.alignment = Alignment(horizontal='right')
        cell_c4.border = thin_border
        
        # Row 5: QI Score with borders
        cell_b5 = ws_trend.cell(row=5, column=2, value='QI Score')
        cell_b5.font = black_font
        cell_b5.border = thin_border
        
        avg_qi = filtered_data['QA Score'].mean() if len(filtered_data) > 0 else 0
        cell_c5 = ws_trend.cell(row=5, column=3, value=round(avg_qi, 2))
        cell_c5.font = black_font
        cell_c5.alignment = Alignment(horizontal='right')
        cell_c5.border = thin_border
        
        # ===== SHEET 4: Team Performance =====
        ws_team = writer.book.create_sheet('Team performance')
        
        # Row 2: Category headers
        c2 = ws_team.cell(row=2, column=3, value='Overall performance')
        c2.font = bold_black_font
        g2 = ws_team.cell(row=2, column=7, value='Introduction & Conclusion')
        g2.font = bold_black_font
        j2 = ws_team.cell(row=2, column=10, value='Problem solving')
        j2.font = bold_black_font
        n2 = ws_team.cell(row=2, column=14, value='Soft skills')
        n2.font = bold_black_font
        
        # Data rows with color coding and borders
        all_metrics_sorted = sorted(
            [(name, perf) for name, perf in metric_performance],
            key=lambda x: x[1],
            reverse=True
        )
        for row_idx, (metric_name, perf) in enumerate(all_metrics_sorted, start=3):
            # Column B: Metric name with border
            cell_b = ws_team.cell(row=row_idx, column=2, value=metric_name)
            cell_b.font = black_font
            cell_b.border = thin_border
            # Determine color based on metric category
            if metric_name in ['Call opening', 'Call closure']:
                cell_b.fill = blue_fill
            elif metric_name in ['Identification of customer needs', 'Educate & Inform', 'Necessary steps to query resolution', 'Initiative', 'Identifying further needs']:
                cell_b.fill = light_green_fill
            else:
                cell_b.fill = gold_fill
            
            # Column C: Performance value with border
            cell_c = ws_team.cell(row=row_idx, column=3, value=perf)
            cell_c.font = black_font
            cell_c.border = thin_border
        
        # Introduction & Conclusion (F3:G) with borders
        intro_data = [
            ('Call opening', metric_performance[0][1]),
            ('Call closure', metric_performance[1][1])
        ]
        for row_idx, (metric_name, perf) in enumerate(intro_data, start=3):
            cell_f = ws_team.cell(row=row_idx, column=6, value=metric_name)
            cell_f.fill = blue_fill
            cell_f.font = black_font
            cell_f.border = thin_border
            cell_g = ws_team.cell(row=row_idx, column=7, value=perf)
            cell_g.font = black_font
            cell_g.border = thin_border
        
        # Problem solving (I3:J) with borders
        problem_data = [
            ('Identification of customer needs', metric_performance[2][1]),
            ('Educate & Inform', metric_performance[3][1]),
            ('Necessary steps to query resolution', metric_performance[4][1]),
            ('Initiative', metric_performance[5][1]),
            ('Identifying further needs', metric_performance[6][1])
        ]
        for row_idx, (metric_name, perf) in enumerate(problem_data, start=3):
            cell_i = ws_team.cell(row=row_idx, column=9, value=metric_name)
            cell_i.fill = light_green_fill
            cell_i.font = black_font
            cell_i.border = thin_border
            cell_j = ws_team.cell(row=row_idx, column=10, value=perf)
            cell_j.font = black_font
            cell_j.border = thin_border
        
        # Soft skills (M3:N) with borders
        soft_data = [
            ('Effective communication', metric_performance[7][1]),
            ('Professionalism', metric_performance[8][1]),
            ('Effective listening & troubleshooting', metric_performance[9][1]),
            ('Politeness & courtesy', metric_performance[10][1]),
            ('Empathy', metric_performance[11][1])
        ]
        for row_idx, (metric_name, perf) in enumerate(soft_data, start=3):
            cell_m = ws_team.cell(row=row_idx, column=13, value=metric_name)
            cell_m.fill = gold_fill
            cell_m.font = black_font
            cell_m.border = thin_border
            cell_n = ws_team.cell(row=row_idx, column=14, value=perf)
            cell_n.font = black_font
            cell_n.border = thin_border
    
    excel_buffer.seek(0)

    st.download_button(
        label="📊 Download Report as Excel",
        data=excel_buffer,
        file_name=f"qa_audit_report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )