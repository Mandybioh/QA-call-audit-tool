import openpyxl
from copy import copy

file_path = r'Score Sheet\Call Centre Audit Report - October 2024 (Recovered).xlsx'
wb = openpyxl.load_workbook(file_path)

# Check Individual Performance sheet merged cells and actual content
ws_ip = wb['Individual performance']
print("=== INDIVIDUAL PERFORMANCE - Merged Cells ===")
print(f"Merged cells: {ws_ip.merged_cells}")
print()

print("=== INDIVIDUAL PERFORMANCE - All cells ===")
for row in ws_ip.iter_rows(min_row=1, max_row=12):
    for cell in row:
        print(f"{cell.coordinate}: '{cell.value}' | Font: {cell.font.name if cell.font else 'None'}, Bold: {cell.font.bold if cell.font else 'None'}, Color: {cell.font.color if cell.font else 'None'} | Fill: {cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'}")
    print()

print()
print("=== TREND SHEET - Merged Cells ===")
ws_trend = wb['Trend']
print(f"Merged cells: {ws_trend.merged_cells}")
print("First 5 rows of content:")
for i, row in enumerate(ws_trend.iter_rows(min_row=1, max_row=5), 1):
    row_str = ""
    for cell in row:
        row_str += f"[{cell.coordinate}: {cell.value}] "
    print(f"Row {i}: {row_str}")
