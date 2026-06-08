import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
import os

# Load template file
template_path = r"Score Sheet\Call Centre Audit Report - October 2024 (Recovered).xlsx"
wb = openpyxl.load_workbook(template_path)

# Get the first sheet
first_sheet = wb.sheetnames[0]
print(f"First sheet name: '{first_sheet}'")
print("=" * 80)

ws = wb[first_sheet]

# Get dimensions
print(f"\nSheet dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

# Analyze row structure
print("\n" + "=" * 80)
print("ROW STRUCTURE ANALYSIS")
print("=" * 80)

# Check first 10 rows for structure
for row_num in range(1, 11):
    row = ws[row_num]
    has_content = False
    for cell in row:
        if cell.value is not None:
            has_content = True
            break
    if has_content:
        print(f"\nRow {row_num}:")
        for col_num in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is not None or cell.fill.start_color.index != '00000000':
                fill_color = cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else 'None'
                font_bold = cell.font.bold if cell.font else False
                font_color = cell.font.color.rgb if (cell.font and hasattr(cell.font.color, 'rgb')) else 'None'
                print(f"  Col {col_num}: Value='{cell.value}' | Color={fill_color} | Bold={font_bold} | FontColor={font_color}")
    else:
        print(f"Row {row_num}: [Empty]")

# Sample data row structure (assuming headers in row 5)
print("\n" + "=" * 80)
print("SAMPLE DATA ROWS")
print("=" * 80)

for row_num in range(6, 12):
    print(f"\nRow {row_num}:")
    for col_num in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value is not None:
            fill_color = cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else 'None'
            alignment = f"h={cell.alignment.horizontal}, v={cell.alignment.vertical}" if cell.alignment else "None"
            print(f"  Col {col_num}: Value='{cell.value}' | Color={fill_color} | Alignment={alignment}")

# Check column widths
print("\n" + "=" * 80)
print("COLUMN WIDTHS")
print("=" * 80)

for col_num in range(1, min(20, ws.max_column + 1)):
    col_letter = openpyxl.utils.get_column_letter(col_num)
    width = ws.column_dimensions[col_letter].width
    print(f"Column {col_letter}: Width={width}")

# Check for merged cells
print("\n" + "=" * 80)
print("MERGED CELLS")
print("=" * 80)
if ws.merged_cells:
    for merged_range in ws.merged_cells.ranges:
        print(f"  {merged_range}")
else:
    print("  No merged cells")

# Check row heights
print("\n" + "=" * 80)
print("ROW HEIGHTS (first 10 rows)")
print("=" * 80)

for row_num in range(1, 11):
    height = ws.row_dimensions[row_num].height
    print(f"Row {row_num}: Height={height}")

print("\n" + "=" * 80)
print("FORMULAS IN FIRST SHEET")
print("=" * 80)

formula_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.data_type == 'f':  # Formula
            print(f"  {cell.coordinate}: {cell.value}")
            formula_count += 1

if formula_count == 0:
    print("  No formulas found in first sheet")

print("\n" + "=" * 80)
print("DETAILED COLOR ANALYSIS (First 20 rows, first 10 columns)")
print("=" * 80)

colors_used = {}
for row_num in range(1, 21):
    for col_num in range(1, 11):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.fill and hasattr(cell.fill.start_color, 'rgb'):
            color = cell.fill.start_color.rgb
            if color and color != '00000000':
                if color not in colors_used:
                    colors_used[color] = []
                colors_used[color].append(f"{cell.coordinate}")

if colors_used:
    print("\nColors found (with sample cells):")
    for color, cells in colors_used.items():
        print(f"  Color {color}: {cells[:3]}...")
else:
    print("  No special colors found")

wb.close()
