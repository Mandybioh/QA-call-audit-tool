from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
import json

wb = load_workbook('Score Sheet/Call Centre Audit Report - October 2024 (Recovered).xlsx', data_only=False)

print("="*100)
print("EXTRACTING FORMULAS, COLORS, AND FORMATTING FROM TEMPLATE")
print("="*100)

# Analyze Oct 2024 sheet
print("\n\n*** OCT 2024 SHEET ***")
ws_oct = wb['Oct 2024']
print(f"Row 3 formulas and values:")
for col in range(1, 28):
    cell = ws_oct.cell(row=3, column=col)
    if cell.value or cell.fill.start_color or cell.font.color:
        fill_color = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill.start_color, 'rgb') else None
        font_color = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None
        font_bold = cell.font.bold if cell.font else None
        print(f"  {cell.coordinate}: value={cell.value}, fill={fill_color}, font_color={font_color}, bold={font_bold}")

# Analyze Trend sheet
print("\n\n*** TREND SHEET - ALL FORMATTING ***")
ws_trend = wb['Trend']
for row in range(1, 7):
    print(f"\nRow {row}:")
    for col in range(1, 19):
        cell = ws_trend.cell(row=row, column=col)
        if cell.value or str(cell.fill.start_color) != '00000000':
            fill_color = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color)
            font_bold = cell.font.bold if cell.font else None
            font_size = cell.font.size if cell.font else None
            font_color = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None
            alignment = cell.alignment.horizontal if cell.alignment else None
            print(f"  {cell.coordinate}: val={cell.value}, fill={fill_color}, bold={font_bold}, size={font_size}, font_color={font_color}, align={alignment}")

# Analyze Team Performance sheet
print("\n\n*** TEAM PERFORMANCE SHEET - ALL FORMATTING ***")
ws_team = wb['Team performance']
for row in range(1, 13):
    print(f"\nRow {row}:")
    for col in range(1, 15):
        cell = ws_team.cell(row=row, column=col)
        if cell.value or str(cell.fill.start_color) != '00000000':
            fill_color = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color)
            font_bold = cell.font.bold if cell.font else None
            font_color = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None
            print(f"  {cell.coordinate}: val={cell.value}, fill={fill_color}, bold={font_bold}, font_color={font_color}")

# Analyze Individual Performance sheet
print("\n\n*** INDIVIDUAL PERFORMANCE SHEET - ALL FORMATTING ***")
ws_ip = wb['Individual performance']
for row in range(1, 8):
    print(f"\nRow {row}:")
    for col in range(1, 4):
        cell = ws_ip.cell(row=row, column=col)
        if cell.value or str(cell.fill.start_color) != '00000000':
            fill_color = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color)
            font_bold = cell.font.bold if cell.font else None
            font_color = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None
            print(f"  {cell.coordinate}: val={cell.value}, fill={fill_color}, bold={font_bold}, font_color={font_color}")

wb.close()
