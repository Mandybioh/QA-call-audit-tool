import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import json

file_path = r'Score Sheet\Call Centre Audit Report - October 2024 (Recovered).xlsx'
wb = openpyxl.load_workbook(file_path)

# Get all sheet names and structure
print("Sheet Names:", wb.sheetnames)
print()

# Examine Individual Performance sheet
ws = wb['Individual performance']
print("=== INDIVIDUAL PERFORMANCE SHEET ===")
print(f"Headers (Row 1):", [cell.value for cell in ws[1]])
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    if i <= 5:
        print(f"Row {i}: {row}")

print()
print("=== TREND SHEET ===")
ws_trend = wb['Trend']
print(f"Content (first 5 rows):")
for i, row in enumerate(ws_trend.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {i}: {row}")

print()
print("=== TEAM PERFORMANCE SHEET ===")
ws_team = wb['Team performance']
print(f"Content (first 10 rows):")
for i, row in enumerate(ws_team.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    print(f"Row {i}: {row}")

print()
print("=== MAIN DATA SHEET (Oct 2024) - HEADERS ===")
ws_main = wb['Oct 2024']
headers = [cell.value for cell in ws_main[1]]
print(f"Total columns: {len(headers)}")
print("Headers:", headers)

print()
print("=== CHECKING COLORS AND FONTS ===")
print("Individual Performance Header Fill:", ws['A1'].fill.start_color.rgb if ws['A1'].fill.start_color else None)
print("Individual Performance Header Font Bold:", ws['A1'].font.bold if ws['A1'].font else None)
print("Individual Performance Header Font Color:", ws['A1'].font.color.rgb if ws['A1'].font and ws['A1'].font.color else None)
